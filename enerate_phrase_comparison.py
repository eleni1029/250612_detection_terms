#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_phrase_comparison.py (v2.2 - Multi-language Version)

根據 i18n_input 中的檔案內容自動生成各語言的 phrase_comparison Excel 檔案
支援多語言敏感詞檢測和自動備份現有檔案

功能：
1. 自動掃描 i18n_input 目錄中的所有語言
2. 檢測每個語言檔案中的敏感詞
3. 為每個語言生成獨立的 phrase_comparison_{language}.xlsx
4. 自動備份現有的 Excel 檔案
"""

import json
import re
import itertools
import sys
import shutil
import datetime
from pathlib import Path
from collections import defaultdict
from config_loader import get_config

try:
    import polib
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"❌ 缺少必要套件：{e}")
    print("請執行：pip install polib openpyxl")
    sys.exit(1)


# 基礎敏感詞字典 - 可根據需要調整
BASE_SENSITIVE_WORDS = {
    "時間相關": [
        "年度", "季度", "月份", "週期", "期間", "日期", "時間", "截止",
        "開始", "結束", "進度", "時程", "排程", "計劃"
    ],
    "數量相關": [
        "總計", "合計", "統計", "數量", "金額", "費用", "成本", "預算",
        "收入", "支出", "利潤", "損失", "餘額", "結餘"
    ],
    "狀態相關": [
        "完成", "進行", "待處理", "已確認", "待確認", "審核", "批准",
        "拒絕", "取消", "暫停", "延期", "終止"
    ],
    "人員相關": [
        "員工", "職員", "主管", "經理", "總監", "客戶", "用戶", "使用者",
        "成員", "參與者", "負責人", "聯絡人"
    ],
    "文件相關": [
        "報告", "文件", "記錄", "檔案", "資料", "信息", "數據", "表單",
        "申請", "提案", "合約", "協議", "證明", "憑證"
    ],
    "業務相關": [
        "項目", "專案", "任務", "工作", "業務", "服務", "產品", "方案",
        "流程", "程序", "標準", "規範", "政策", "制度"
    ]
}


def main():
    """主執行函數"""
    print("🚀 開始生成多語言 phrase_comparison Excel 檔案")
    
    # 載入配置
    config = get_config()
    config.print_config_summary()
    
    # 檢測可用語言
    available_languages = config.detect_available_languages()
    print(f"\n🌐 將處理 {len(available_languages)} 個語言：{', '.join(available_languages)}")
    
    # 備份配置
    backup_config = config.get_backup_config()
    backup_dir = config.get_backup_dir()
    backup_dir.mkdir(exist_ok=True)
    
    timestamp_format = backup_config.get('timestamp_format', '%Y%m%d_%H%M%S')
    timestamp = datetime.datetime.now().strftime(timestamp_format)
    
    # 處理每個語言
    for language in available_languages:
        print(f"\n📋 處理語言：{language}")
        process_language(config, language, timestamp)
    
    print(f"\n🎉 所有語言的 phrase_comparison 檔案生成完成！")
    print(f"📄 備份檔案位於：{backup_dir}")


def process_language(config, language: str, timestamp: str):
    """
    處理單個語言的 phrase_comparison 生成
    
    Args:
        config: 配置物件
        language: 語言代碼
        timestamp: 時間戳
    """
    
    # 獲取檔案路徑
    excel_path = config.get_comparison_excel_path(language)
    language_files = config.get_language_files(language)
    
    print(f"   目標檔案：{excel_path}")
    print(f"   來源檔案：{list(language_files.values())}")
    
    # 備份現有檔案
    if excel_path.exists():
        backup_dir = config.get_backup_dir()
        backup_filename = f"{excel_path.stem}_{timestamp}{excel_path.suffix}"
        backup_path = backup_dir / backup_filename
        
        shutil.copy2(excel_path, backup_path)
        print(f"   ✅ 已備份現有檔案：{backup_filename}")
    
    # 檢測敏感詞
    detected_keywords = detect_sensitive_words(language_files, config)
    
    if not detected_keywords:
        print(f"   ⚠️  在 {language} 中未檢測到敏感詞，使用基礎詞彙")
        detected_keywords = BASE_SENSITIVE_WORDS.copy()
    
    print(f"   📊 檢測到 {sum(len(words) for words in detected_keywords.values())} 個敏感詞")
    for category, words in detected_keywords.items():
        print(f"      {category}: {len(words)} 個詞")
    
    # 生成 Excel
    generate_comparison_excel(config, language, detected_keywords, excel_path)
    print(f"   ✅ 生成完成：{excel_path}")


def detect_sensitive_words(language_files: dict, config) -> dict:
    """
    從語言檔案中檢測敏感詞
    
    Args:
        language_files: 語言檔案路徑字典
        config: 配置物件
    
    Returns:
        檢測到的敏感詞字典 {分類: [詞彙列表]}
    """
    
    # 收集所有文本內容
    all_texts = []
    
    # 讀取 PO 檔案
    if 'po_file' in language_files:
        try:
            po_file = polib.pofile(str(language_files['po_file']))
            for entry in po_file:
                if entry.msgid:
                    all_texts.append(entry.msgid)
                if entry.msgstr:
                    all_texts.append(entry.msgstr)
        except Exception as e:
            print(f"      ⚠️  讀取 PO 檔案失敗：{e}")
    
    # 讀取 JSON 檔案
    if 'json_file' in language_files:
        try:
            with open(language_files['json_file'], 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            def extract_json_values(obj):
                """遞迴提取 JSON 中的所有字符串值"""
                if isinstance(obj, dict):
                    for value in obj.values():
                        yield from extract_json_values(value)
                elif isinstance(obj, list):
                    for item in obj:
                        yield from extract_json_values(item)
                elif isinstance(obj, str):
                    yield obj
            
            all_texts.extend(extract_json_values(json_data))
            
        except Exception as e:
            print(f"      ⚠️  讀取 JSON 檔案失敗：{e}")
    
    if not all_texts:
        print(f"      ⚠️  無法從檔案中提取文本內容")
        return {}
    
    # 檢測敏感詞
    print(f"      🔍 從 {len(all_texts)} 個文本條目中檢測敏感詞...")
    
    detected_words = defaultdict(set)
    detection_config = config.get_keyword_detection_config()
    case_sensitive = detection_config.get('case_sensitive', False)
    
    # 合併所有文本
    combined_text = ' '.join(all_texts)
    if not case_sensitive:
        combined_text = combined_text.lower()
    
    # 對每個基礎分類的敏感詞進行檢測
    for category, base_words in BASE_SENSITIVE_WORDS.items():
        for word in base_words:
            search_word = word.lower() if not case_sensitive else word
            if search_word in combined_text:
                detected_words[category].add(word)
    
    # 轉換為普通字典，並按原始順序排列
    result = {}
    for category, words in detected_words.items():
        # 保持與基礎詞典相同的順序
        ordered_words = []
        for base_word in BASE_SENSITIVE_WORDS[category]:
            if base_word in words:
                ordered_words.append(base_word)
        if ordered_words:
            result[category] = ordered_words
    
    return result


def generate_comparison_excel(config, language: str, keywords_dict: dict, output_path: Path):
    """
    生成 phrase_comparison Excel 檔案
    
    Args:
        config: 配置物件
        language: 語言代碼
        keywords_dict: 敏感詞字典
        output_path: 輸出檔案路徑
    """
    
    # 創建工作簿
    wb = Workbook()
    ws = wb.active
    
    # 設置工作表名稱
    excel_config = config.get_excel_config()
    ws.title = excel_config.get('worksheet_name', 'phrase_comparison')
    
    # 建立標題列
    headers = ["敏感詞類型", "敏感詞"]
    
    # 添加各業態的對應方案欄位
    business_types = config.get_business_types()
    business_columns = excel_config.get('business_columns', {})
    solution_template = business_columns.get('solution_template', '對應方案({display_name})')
    
    for bt_code, bt_config in business_types.items():
        display_name = bt_config['display_name']
        column_name = solution_template.format(display_name=display_name)
        headers.append(column_name)
    
    # 寫入標題列
    ws.append(headers)
    
    # 設置標題樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 寫入資料列
    row_num = 2
    total_keywords = 0
    
    for category, keywords in keywords_dict.items():
        for keyword in keywords:
            row_data = [category, keyword]
            
            # 為每個業態添加空白的對應方案欄位（讓使用者填寫）
            for bt_code in business_types.keys():
                row_data.append("")  # 空白，讓使用者手動填寫
            
            ws.append(row_data)
            total_keywords += 1
            row_num += 1
    
    # 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        
        for cell in col:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # 設定欄寬，最小15，最大50
        adjusted_width = min(max(max_length + 4, 15), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 設置數據區域的對齊方式
    data_alignment = Alignment(horizontal="left", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=row_num-1):
        for cell in row:
            cell.alignment = data_alignment
    
    # 確保輸出目錄存在
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 保存檔案
    wb.save(output_path)
    
    print(f"      📊 Excel 統計：{total_keywords} 個敏感詞，{len(business_types)} 個業態欄位")


def test_detection():
    """測試敏感詞檢測功能"""
    print("🧪 測試敏感詞檢測功能...")
    
    # 創建測試文本
    test_texts = [
        "年度報告顯示總計金額達到預期",
        "季度數據統計員工完成率",
        "項目進度報告需要主管審核",
        "客戶文件記錄已確認無誤"
    ]
    
    # 模擬檢測
    combined_text = ' '.join(test_texts).lower()
    
    detected = defaultdict(list)
    for category, words in BASE_SENSITIVE_WORDS.items():
        for word in words:
            if word in combined_text:
                detected[category].append(word)
    
    print("檢測結果：")
    for category, words in detected.items():
        print(f"  {category}: {', '.join(words)}")
    
    return dict(detected)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='生成多語言 phrase_comparison Excel 檔案')
    parser.add_argument('--test', action='store_true', help='執行檢測測試')
    parser.add_argument('--language', '-l', help='只處理指定語言')
    
    args = parser.parse_args()
    
    if args.test:
        test_detection()
    else:
        if args.language:
            # 處理單一語言
            config = get_config()
            available_languages = config.detect_available_languages()
            
            if args.language not in available_languages:
                print(f"❌ 語言 '{args.language}' 不在可用列表中：{available_languages}")
                sys.exit(1)
            
            print(f"🚀 處理單一語言：{args.language}")
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            process_language(config, args.language, timestamp)
            print(f"✅ 完成")
        else:
            # 處理所有語言
            main()