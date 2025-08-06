#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
script_02_apply_combine.py (v1.7 - 修復空檔案生成版)

功能：
1. 選擇要合併的 tobemodified Excel 檔案（支援多選）
2. 選擇 i18n_combine 目錄下的 JSON/PO 檔案作為合併目標
3. 按業態分別處理，避免相互衝突
4. 沒有目標檔案時自動創建標準檔案（JSON/PO）
5. 生成合併後的檔案到 i18n_output/multi_{timestamp}_combined/
6. 提供詳細的合併報告和日誌
7. **完整陣列更新邏輯 - 從 i18n_input 讀取原始陣列進行智能合併**
8. **v1.7 新增：智能檔案生成 - 只在有實際內容時才生成 JSON 檔案，避免空檔案**
"""

import json
import sys
import shutil
import datetime
import argparse
import glob
from pathlib import Path
from collections import defaultdict
from config_loader import get_config

try:
    import openpyxl
    import polib
except ImportError as e:
    print(f"❌ 缺少必要套件：{e}")
    print("請執行：pip install openpyxl polib")
    sys.exit(1)


def check_multilang_json_structure(data: dict) -> bool:
    """檢查 JSON 是否為多語言結構（簡化版）"""
    if not isinstance(data, dict):
        return False
    
    # 簡化的檢查：如果頂層 key 看起來像語言代碼（2-5個字符），則認為是多語言結構
    for key in data.keys():
        if isinstance(key, str) and 2 <= len(key) <= 10 and isinstance(data[key], dict):
            return True
    
    return False


def load_original_language_json(language: str) -> dict:
    """載入指定語言的原始 JSON 檔案 (i18n_input/{language}/{language}.json)"""
    try:
        input_dir = Path("i18n_input")
        language_file = input_dir / language / f"{language}.json"
        
        if not language_file.exists():
            print(f"⚠️  原始語言檔案不存在：{language_file}")
            return {}
        
        with open(language_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            print(f"✅ 載入原始語言檔案：{language_file}")
            return data
    
    except Exception as e:
        print(f"❌ 載入原始語言檔案失敗 ({language})：{e}")
        return {}


def detect_array_path_and_index(path: str) -> tuple:
    """
    檢測路徑是否包含陣列索引，並返回陣列路徑和索引
    
    Returns:
        (array_path, index) 如果是陣列索引路徑
        (None, None) 如果不是陣列索引路徑
    
    例如：
        "slogan[1]" -> ("slogan", 1)
        "data.items[0].tags[2]" -> ("data.items[0].tags", 2)
        "simple.key" -> (None, None)
    """
    import re
    
    # 使用正規表達式找到最後一個陣列索引
    pattern = r'^(.+)\[(\d+)\]$'
    match = re.match(pattern, path)
    
    if match:
        array_path = match.group(1)
        index = int(match.group(2))
        return (array_path, index)
    
    return (None, None)


def get_array_from_original_json(original_data: dict, array_path: str) -> list:
    """從原始 JSON 資料中獲取指定路徑的陣列"""
    try:
        path_parts = parse_json_path(array_path)
        current = original_data
        
        for part_type, part_value in path_parts:
            if part_type == 'key':
                if part_value not in current:
                    print(f"⚠️  原始資料中找不到路徑：{array_path}")
                    return []
                current = current[part_value]
            elif part_type == 'index':
                if not isinstance(current, list) or len(current) <= part_value:
                    print(f"⚠️  原始資料中陣列索引超出範圍：{array_path}")
                    return []
                current = current[part_value]
        
        if isinstance(current, list):
            return current.copy()  # 返回副本避免修改原始資料
        else:
            print(f"⚠️  指定路徑不是陣列：{array_path} (類型: {type(current)})")
            return []
            
    except Exception as e:
        print(f"❌ 從原始資料獲取陣列失敗：{array_path} - {e}")
        return []


def create_default_json_file(output_path: Path, all_updates: dict, detected_languages: list) -> bool:
    """創建預設的多語言 JSON 檔案（僅包含檢測到的語言區塊）"""
    try:
        # 根據檢測到的語言建立空結構
        json_data = {}
        
        # 只添加檢測到的語言，創建空結構
        for language in detected_languages:
            json_data[language] = {}
        
        # 根據 Excel 更新資料動態添加路徑結構（但不設置值）
        for language, language_updates in all_updates.items():
            if language not in json_data:
                json_data[language] = {}
                
            for bt_code, bt_updates in language_updates.items():
                for json_path_str, new_value, update_language in bt_updates['json']:
                    # 確保路徑存在於對應語言中
                    if update_language in json_data:
                        # 預先創建路徑結構，但不設置值（將由後續合併處理）
                        create_json_path_structure(json_data[update_language], json_path_str)
        
        # 確保輸出目錄存在
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 保存檔案
        json_content = json.dumps(json_data, ensure_ascii=False, indent=2)
        output_path.write_text(json_content, encoding="utf-8")
        
        return True
        
    except Exception as e:
        print(f"❌ 創建預設 JSON 檔案失敗：{e}")
        return False


def create_json_path_structure(data: dict, path: str):
    """在 JSON 中預先創建路徑結構"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for i, (part_type, part_value) in enumerate(path_parts):
            is_last = (i == len(path_parts) - 1)
            
            if part_type == 'key':
                if not is_last:
                    if part_value not in current:
                        # 檢查下一部分是否為索引
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
                else:
                    # 最後一個部分，如果不存在則設為空字串
                    if part_value not in current:
                        current[part_value] = ""
            
            elif part_type == 'index':
                if not is_last:
                    while len(current) <= part_value:
                        current.append({})
                    current = current[part_value]
                else:
                    while len(current) <= part_value:
                        current.append("")
        
    except Exception as e:
        print(f"⚠️  創建JSON路徑結構失敗：{path} - {e}")


def create_default_po_file(output_path: Path, language: str = "zh_Hant_TW") -> bool:
    """創建預設的 messages.po 檔案（僅包含標頭，無範例條目）"""
    try:
        # 創建新的 PO 檔案
        po = polib.POFile()
        
        # 設置標頭資訊
        current_time = datetime.datetime.now()
        po.metadata = {
            'Project-Id-Version': 'PROJECT VERSION',
            'Report-Msgid-Bugs-To': 'EMAIL@ADDRESS',
            'POT-Creation-Date': current_time.strftime('%Y-%m-%d %H:%M%z'),
            'PO-Revision-Date': 'YEAR-MO-DA HO:MI+ZONE',
            'Last-Translator': 'FULL NAME <EMAIL@ADDRESS>',
            'Language': language,
            'Language-Team': f'{language} <LL@li.org>',
            'Plural-Forms': 'nplurals=1; plural=0;',
            'MIME-Version': '1.0',
            'Content-Type': 'text/plain; charset=utf-8',
            'Content-Transfer-Encoding': '8bit',
            'Generated-By': 'Babel 2.12.1'
        }
        
        # 確保輸出目錄存在
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 保存檔案
        po.save(str(output_path))
        
        return True
        
    except Exception as e:
        print(f"❌ 創建預設 PO 檔案失敗：{e}")
        return False


def detect_tobemodified_files(config) -> dict:
    """檢測可用的 tobemodified 檔案"""
    available_files = {}
    
    # 檢測輸出目錄中的檔案
    try:
        dirs = config.get_directories()
        output_dir = Path(dirs['output_dir'])
    except Exception:
        output_dir = Path('i18n_output')
    
    # 使用配置載入器的語言檢測
    try:
        available_languages = config.detect_available_languages()
    except Exception as e:
        print(f"⚠️  語言檢測失敗：{e}")
        available_languages = []
    
    # 檢測所有帶時間戳的 tobemodified 檔案
    import re
    for language in available_languages:
        # 尋找該語言的所有 tobemodified 檔案（包括帶時間戳的）
        pattern = f"{language}_tobemodified*.xlsx"
        language_files = list(output_dir.glob(pattern))
        
        if language_files:
            # 按檔案修改時間排序（從新到舊）
            language_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            available_files[language] = language_files
    
    return available_files


def scan_combine_directory(combine_dir: Path) -> dict:
    """掃描 i18n_combine 目錄中的檔案"""
    files = {
        'json': [],
        'po': []
    }
    
    if not combine_dir.exists():
        return files
    
    # 遞歸掃描所有 JSON 和 PO 檔案
    for file_path in combine_dir.rglob("*.json"):
        relative_path = file_path.relative_to(combine_dir)
        files['json'].append({
            'path': file_path,
            'relative_path': str(relative_path),
            'name': file_path.name
        })
    
    for file_path in combine_dir.rglob("*.po"):
        relative_path = file_path.relative_to(combine_dir)
        files['po'].append({
            'path': file_path,
            'relative_path': str(relative_path),
            'name': file_path.name
        })
    
    return files


def choose_tobemodified_files(available_files: dict) -> dict:
    """選擇要使用的 tobemodified 檔案（支援多選和多版本選擇）"""
    if not available_files:
        print("❌ 未找到任何 tobemodified 檔案")
        return {}
    
    # 展開所有語言的檔案選項
    all_choices = []
    for language, file_list in available_files.items():
        if isinstance(file_list, list):
            for i, file_path in enumerate(file_list):
                # 顯示時間戳和檔案大小等資訊
                import datetime
                mtime = datetime.datetime.fromtimestamp(file_path.stat().st_mtime)
                time_str = mtime.strftime('%Y-%m-%d %H:%M:%S')
                all_choices.append((language, file_path, time_str, i == 0))  # 第一個是最新的
        else:
            # 兼容舊格式
            file_path = file_list
            mtime = datetime.datetime.fromtimestamp(file_path.stat().st_mtime)
            time_str = mtime.strftime('%Y-%m-%d %H:%M:%S')
            all_choices.append((language, file_path, time_str, True))
    
    if not all_choices:
        print("❌ 沒有找到有效的 tobemodified 檔案")
        return {}
    
    print("\n📄 可用的 tobemodified 檔案（按時間排序，從新到舊）：")
    
    # 按語言分組顯示
    language_groups = {}
    for language, file_path, time_str, is_latest in all_choices:
        if language not in language_groups:
            language_groups[language] = []
        language_groups[language].append((file_path, time_str, is_latest))
    
    choice_index = 1
    indexed_choices = []
    
    for language in sorted(language_groups.keys()):
        print(f"\n  🌐 {language}:")
        for file_path, time_str, is_latest in language_groups[language]:
            latest_mark = " [最新]" if is_latest else ""
            print(f"    {choice_index}) {file_path.name} ({time_str}){latest_mark}")
            indexed_choices.append((language, file_path))
            choice_index += 1
    
    print(f"\n  L) 每個語言自動選擇最新版本")
    print(f"  A) 全部選擇")
    print(f"  0) 取消操作")
    
    selected_files = {}
    
    while True:
        try:
            choice = input(f"\n請選擇要使用的檔案 (可多選，用逗號分隔，如 1,2,3 或 L/A)：").strip()
            
            if choice == '0':
                print("❌ 操作取消")
                return {}
            elif choice.upper() == 'L':
                # 每個語言選擇最新版本
                for language, file_list in available_files.items():
                    if isinstance(file_list, list):
                        selected_files[language] = file_list[0]  # 第一個是最新的
                    else:
                        selected_files[language] = file_list
                break
            elif choice.upper() == 'A':
                # 選擇所有檔案（每個語言的所有版本中選最新的）
                for language, file_list in available_files.items():
                    if isinstance(file_list, list):
                        selected_files[language] = file_list[0]
                    else:
                        selected_files[language] = file_list
                break
            else:
                # 解析多選
                choice_indices = [int(x.strip()) - 1 for x in choice.split(',')]
                selected_files = {}
                
                for choice_idx in choice_indices:
                    if 0 <= choice_idx < len(indexed_choices):
                        language, file_path = indexed_choices[choice_idx]
                        selected_files[language] = file_path
                    else:
                        print(f"⚠️  無效選項：{choice_idx + 1}")
                        continue
                
                if selected_files:
                    break
                else:
                    print(f"⚠️  請輸入有效的選項")
                    
        except (ValueError, KeyboardInterrupt):
            print("\n❌ 操作取消")
            return {}
    
    print(f"✅ 選擇了 {len(selected_files)} 個檔案：")
    for language, file_path in selected_files.items():
        mtime = datetime.datetime.fromtimestamp(file_path.stat().st_mtime)
        time_str = mtime.strftime('%Y-%m-%d %H:%M:%S')
        print(f"   {language}: {file_path.name} ({time_str})")
    
    return selected_files


def choose_combine_file(files: list, file_type: str) -> Path:
    """選擇要合併的檔案"""
    if not files:
        print(f"⚠️  /i18n_combine/ 中沒有找到 {file_type.upper()} 檔案")
        if file_type.lower() == 'po':
            print(f"💡 將自動創建預設的 messages.po 檔案")
        elif file_type.lower() == 'json':
            print(f"💡 將自動創建預設的多語言 JSON 檔案")
        return None
    
    print(f"\n📁 可用的 {file_type.upper()} 檔案：")
    for i, file_info in enumerate(files, 1):
        print(f"  {i}) {file_info['relative_path']}")
    
    print(f"  0) 跳過 {file_type.upper()} 檔案")
    if file_type.lower() in ['po', 'json']:
        create_option = "messages.po" if file_type.lower() == 'po' else "多語言 JSON"
        print(f"  C) 創建新的 {create_option} 檔案")
    
    while True:
        try:
            choice = input(f"\n請選擇要合併的 {file_type.upper()} 檔案 (0-{len(files)}{'/C' if file_type.lower() in ['po', 'json'] else ''})：").strip()
            
            if choice == '0':
                print(f"⏭️  跳過 {file_type.upper()} 檔案")
                return None
            elif choice.upper() == 'C' and file_type.lower() in ['po', 'json']:
                create_option = "messages.po" if file_type.lower() == 'po' else "多語言 JSON"
                print(f"🆕 將創建新的 {create_option} 檔案")
                return "CREATE_NEW"
            else:
                choice_idx = int(choice)
                if 1 <= choice_idx <= len(files):
                    selected_file = files[choice_idx - 1]
                    print(f"✅ 選擇了：{selected_file['relative_path']}")
                    return selected_file['path']
                else:
                    suffix = ' 或 C' if file_type.lower() in ['po', 'json'] else ''
                    print(f"⚠️  請輸入 0-{len(files)} 之間的數字{suffix}")
        except (ValueError, KeyboardInterrupt):
            print("\n❌ 操作取消")
            return None


def read_excel_updates_for_language(xlsx_path: Path, language: str, config) -> dict:
    """讀取單個語言的 Excel 檔案中的更新資料"""
    try:
        print(f"📖 讀取 {language} 的 Excel 檔案：{xlsx_path.name}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        header_row = list(ws[1])
        header = {cell.value: idx for idx, cell in enumerate(header_row) if cell.value}
        
        # 基本欄位檢查
        required_columns = ["檔案類型", "項目ID", "項目內容"]
        missing_columns = []
        
        for col in required_columns:
            if col not in header:
                missing_columns.append(col)
        
        if missing_columns:
            print(f"❌ {language} Excel 缺少必要欄位：{missing_columns}")
            return {}
        
        # 自動檢測所有業態的替換結果欄位
        business_types = config.get_business_types()
        available_business_types = []
        
        for bt_code, bt_config in business_types.items():
            display_name = bt_config['display_name']
            result_col_name = f"{display_name}_替換結果"
            if result_col_name in header:
                available_business_types.append(bt_code)
        
        if not available_business_types:
            print(f"❌ {language} 未找到任何業態的替換結果欄位")
            return {}
        
        print(f"   📋 {language} 檢測到業態：{', '.join([business_types[bt]['display_name'] for bt in available_business_types])}")
        
        # 解析更新資料
        updates = {bt_code: {"po": [], "json": []} for bt_code in available_business_types}
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= max(header.values()):
                continue
            
            try:
                file_type = row[header["檔案類型"]]
                entry_id = row[header["項目ID"]]
                original_text = row[header["項目內容"]]
                
                if not file_type or not entry_id:
                    continue
                
                file_type = str(file_type).lower()
                
                # 處理每個可用的業態
                for bt_code in available_business_types:
                    display_name = business_types[bt_code]['display_name']
                    result_col_name = f"{display_name}_替換結果"
                    
                    new_value = row[header[result_col_name]]
                    
                    # 跳過空值和與原文相同的值
                    if not new_value or not str(new_value).strip():
                        continue
                    
                    new_value = str(new_value).strip()
                    
                    if original_text and str(original_text).strip() == new_value:
                        continue
                    
                    # 創建更新記錄，包含語言信息
                    update_record = (str(entry_id), new_value, language)
                    
                    if file_type == "po" or file_type == "combine_po":
                        updates[bt_code]["po"].append(update_record)
                    elif file_type == "json" or file_type == "combine_json":
                        updates[bt_code]["json"].append(update_record)
            
            except Exception as e:
                print(f"⚠️  {language} 第 {row_num} 行處理失敗: {e}")
                continue
        
        # 統計有效更新
        total_updates = 0
        for bt_code in available_business_types:
            bt_updates = len(updates[bt_code]["po"]) + len(updates[bt_code]["json"])
            total_updates += bt_updates
            if bt_updates > 0:
                print(f"     {business_types[bt_code]['display_name']}: {bt_updates} 個更新")
        
        print(f"   📊 {language} 總計：{total_updates} 個有效更新")
        return updates
        
    except Exception as e:
        print(f"❌ 讀取 {language} Excel 檔案失敗：{e}")
        return {}


def has_non_empty_content(obj) -> bool:
    """【v1.7 新增】檢查物件是否包含非空內容"""
    if isinstance(obj, dict):
        return any(has_non_empty_content(v) for v in obj.values())
    elif isinstance(obj, list):
        return any(has_non_empty_content(item) for item in obj)
    elif isinstance(obj, str):
        return bool(obj.strip())
    else:
        return obj is not None and obj != ""


def combine_multilang_json_files_for_business_type(all_updates: dict, target_json_path: Path, 
                                                  output_json_path: Path, bt_code: str, log_detail=None,
                                                  create_new: bool = False, detected_languages: list = None) -> dict:
    """【v1.7 增強版】為特定業態合併多語言 JSON 檔案，支援完整陣列更新，只在有實際內容時才生成檔案"""
    result = {
        "success": False,
        "merged": 0,
        "skipped": 0,
        "conflicts": [],
        "errors": [],
        "language_stats": {},
        "created_new": False,
        "file_skipped": False  # 【v1.7 新增】檔案是否被跳過
    }
    
    # 檢查是否有當前業態的更新
    json_updates_for_bt = []
    for language_updates in all_updates.values():
        if bt_code in language_updates and language_updates[bt_code]['json']:
            json_updates_for_bt.extend(language_updates[bt_code]['json'])
    
    if not json_updates_for_bt and not create_new:
        result["success"] = True
        if log_detail:
            log_detail(f"JSON ({bt_code}): 沒有任何更新項目")
        return result
    
    try:
        # 處理目標 JSON 檔案
        is_creating_new_file = False  # 新增標記變數
        
        if create_new or target_json_path == "CREATE_NEW" or not target_json_path or not target_json_path.exists():
            # 創建新的 JSON 檔案
            print(f"   🆕 創建新的多語言 JSON 檔案：{output_json_path.name}")
            if log_detail:
                log_detail(f"創建新的多語言 JSON 檔案：{output_json_path.name}")
            
            # 創建預設檔案到臨時位置
            temp_json_path = output_json_path.parent / f"temp_multilang.json"
            temp_json_path.parent.mkdir(parents=True, exist_ok=True)
            
            if not create_default_json_file(temp_json_path, all_updates, detected_languages or []):
                result["errors"].append(f"無法創建預設 JSON 檔案")
                return result
            
            target_data = json.loads(temp_json_path.read_text(encoding="utf-8"))
            result["created_new"] = True
            is_creating_new_file = True  # 設置為新建檔案標記
            
        else:
            # 載入現有的 JSON 檔案
            target_data = json.loads(target_json_path.read_text(encoding="utf-8"))
            print(f"   📄 載入目標多語言 JSON 檔案：{target_json_path.name}")
            if log_detail:
                log_detail(f"載入目標 JSON 檔案：{target_json_path.name}")
            is_creating_new_file = False  # 明確設置為非新建檔案
        
        # 檢查是否為多語言結構
        is_multilang_structure = check_multilang_json_structure(target_data)
        print(f"   🔍 多語言結構檢測：{'是' if is_multilang_structure else '否'}")
        if log_detail:
            log_detail(f"多語言結構檢測：{'是' if is_multilang_structure else '否'}")
        
        # 載入所有語言的原始資料用於陣列更新
        original_language_data = {}
        for language in all_updates.keys():
            original_data = load_original_language_json(language)
            if original_data:
                original_language_data[language] = original_data
                if log_detail:
                    log_detail(f"載入 {language} 原始資料用於陣列更新")
        
        conflicts = []
        language_stats = {}
        
        # 只處理當前業態的更新
        for language, language_updates in all_updates.items():
            if bt_code not in language_updates:
                continue
                
            language_stats[language] = {"merged": 0, "skipped": 0, "conflicts": 0}
            
            if log_detail:
                log_detail(f"處理語言 {language} 的 JSON 更新 (業態: {bt_code})")
            
            # 處理當前業態的 JSON 更新
            bt_updates = language_updates[bt_code]
            for json_path_str, new_value, update_language in bt_updates['json']:
                if log_detail:
                    log_detail(f"處理更新：{update_language}.{json_path_str} = {new_value}")
                
                # 檢測是否為陣列索引路徑
                array_path, array_index = detect_array_path_and_index(json_path_str)
                
                if array_path is not None and array_index is not None:
                    # 這是陣列索引更新，需要進行完整陣列更新
                    if log_detail:
                        log_detail(f"檢測到陣列索引更新：{array_path}[{array_index}] = {new_value}")
                    
                    # 從原始語言資料中獲取完整陣列
                    if update_language in original_language_data:
                        original_array = get_array_from_original_json(original_language_data[update_language], array_path)
                        
                        if original_array:
                            # 確保陣列足夠長
                            while len(original_array) <= array_index:
                                original_array.append("")
                            
                            # 更新指定索引的值
                            original_array[array_index] = new_value
                            
                            # 多語言結構的路徑映射
                            if is_multilang_structure:
                                final_path = f"{update_language}.{array_path}"
                            else:
                                final_path = array_path
                            
                            # 設置完整陣列到目標路徑
                            if set_json_value_by_path(target_data, final_path, original_array):
                                result["merged"] += 1
                                language_stats[update_language]["merged"] += 1
                                if log_detail:
                                    log_detail(f"完整陣列更新成功：{final_path} = {original_array}")
                            else:
                                error_msg = f"無法設置完整陣列：{final_path}"
                                result["errors"].append(error_msg)
                                if log_detail:
                                    log_detail(f"錯誤：{error_msg}")
                        else:
                            # 無法獲取原始陣列，使用傳統方式
                            if log_detail:
                                log_detail(f"無法獲取原始陣列，使用傳統索引更新：{json_path_str}")
                            
                            # 多語言結構的路徑映射
                            if is_multilang_structure:
                                multilang_path = f"{update_language}.{json_path_str}"
                            else:
                                multilang_path = json_path_str
                            
                            # 傳統的索引更新方式
                            if set_json_value_by_path(target_data, multilang_path, new_value):
                                result["merged"] += 1
                                language_stats[update_language]["merged"] += 1
                                if log_detail:
                                    log_detail(f"傳統索引更新成功：{multilang_path} = {new_value}")
                            else:
                                error_msg = f"無法設置傳統索引路徑：{multilang_path}"
                                result["errors"].append(error_msg)
                                if log_detail:
                                    log_detail(f"錯誤：{error_msg}")
                    else:
                        if log_detail:
                            log_detail(f"未找到 {update_language} 的原始資料，使用傳統更新方式")
                        
                        # 多語言結構的路徑映射
                        if is_multilang_structure:
                            multilang_path = f"{update_language}.{json_path_str}"
                        else:
                            multilang_path = json_path_str
                        
                        # 傳統的索引更新方式
                        if set_json_value_by_path(target_data, multilang_path, new_value):
                            result["merged"] += 1
                            language_stats[update_language]["merged"] += 1
                            if log_detail:
                                log_detail(f"傳統索引更新成功：{multilang_path} = {new_value}")
                        else:
                            error_msg = f"無法設置傳統索引路徑：{multilang_path}"
                            result["errors"].append(error_msg)
                            if log_detail:
                                log_detail(f"錯誤：{error_msg}")
                
                else:
                    # 這是普通路徑更新（非陣列索引）
                    # 多語言結構的路徑映射
                    if is_multilang_structure:
                        multilang_path = f"{update_language}.{json_path_str}"
                    else:
                        multilang_path = json_path_str
                    
                    # 獲取現有值
                    existing_value = get_json_value_by_path(target_data, multilang_path)
                    
                    # 修正的衝突檢測邏輯：新建檔案時跳過衝突檢測
                    if not is_creating_new_file and existing_value is not None:
                        existing_str = str(existing_value).strip()
                        new_str = str(new_value).strip()
                        
                        # 如果值完全相同，跳過
                        if existing_str == new_str:
                            result["skipped"] += 1
                            language_stats[update_language]["skipped"] += 1
                            if log_detail:
                                log_detail(f"跳過相同值：{multilang_path} = '{new_str}'")
                            continue
                        
                        # 當值不同且不是空字串時，標記為衝突並讓用戶決定
                        if existing_str != new_str and existing_str != "":
                            conflict_info = {
                                "path": multilang_path,
                                "language": update_language,
                                "existing_value": existing_str,
                                "new_value": new_str,
                                "file_type": "json"
                            }
                            conflicts.append(conflict_info)
                            result["conflicts"].append(conflict_info)
                            language_stats[update_language]["conflicts"] += 1
                            
                            if log_detail:
                                log_detail(f"發現衝突：{multilang_path}")
                                log_detail(f"  現有值: '{existing_str}'")
                                log_detail(f"  新值: '{new_str}'")
                            
                            # 詢問用戶如何處理衝突
                            choice = handle_json_conflict(multilang_path, existing_str, new_str, update_language)
                            
                            if choice == "keep_existing":
                                result["skipped"] += 1
                                language_stats[update_language]["skipped"] += 1
                                if log_detail:
                                    log_detail(f"保留現有值：{multilang_path} = '{existing_str}'")
                                continue
                            elif choice == "use_new":
                                # 繼續執行更新邏輯
                                if log_detail:
                                    log_detail(f"採用新值：{multilang_path} = '{new_str}'")
                            elif choice == "skip":
                                result["skipped"] += 1
                                language_stats[update_language]["skipped"] += 1
                                if log_detail:
                                    log_detail(f"跳過處理：{multilang_path}")
                                continue
                    
                    # 應用普通更新
                    if set_json_value_by_path(target_data, multilang_path, new_value):
                        result["merged"] += 1
                        language_stats[update_language]["merged"] += 1
                        if log_detail:
                            if is_creating_new_file:
                                log_detail(f"新建檔案寫入：{multilang_path} = '{new_value}'")
                            else:
                                original_display = f"'{existing_value}'" if existing_value is not None else "無"
                                log_detail(f"成功更新：{multilang_path} = '{new_value}' (原值: {original_display})")
                    else:
                        error_msg = f"無法設置 JSON 路徑：{multilang_path} (語言: {update_language})"
                        result["errors"].append(error_msg)
                        if log_detail:
                            log_detail(f"錯誤：{error_msg}")
        
        # 【v1.7 修改】智能檔案保存邏輯 - 只在有實際內容時才保存檔案
        should_save_file = False
        
        # 檢查是否有實際更新內容
        if result["merged"] > 0:
            should_save_file = True
            if log_detail:
                log_detail(f"檢測到 {result['merged']} 個實際更新，將保存檔案")
        
        # 如果是創建新檔案，檢查是否有非空內容
        elif result["created_new"]:
            # 檢查 target_data 是否包含非空內容
            has_content = has_non_empty_content(target_data)
            
            if has_content:
                should_save_file = True
                if log_detail:
                    log_detail("新檔案包含實際內容，將保存檔案")
            else:
                if log_detail:
                    log_detail("新檔案無實際內容，跳過保存")
        
        if should_save_file:
            # 保存合併後的檔案
            output_json_path.parent.mkdir(parents=True, exist_ok=True)
            
            json_content = json.dumps(target_data, ensure_ascii=False, indent=2)
            output_json_path.write_text(json_content, encoding="utf-8")
            
            if log_detail:
                log_detail(f"JSON 檔案已保存：{output_json_path}")
        else:
            if log_detail:
                log_detail(f"跳過保存空的 JSON 檔案：{output_json_path}")
            
            # 設置特殊標記表示檔案未保存
            result["file_skipped"] = True
        
        # 清理臨時檔案
        temp_json_path = output_json_path.parent / f"temp_multilang.json"
        if temp_json_path.exists():
            temp_json_path.unlink()
        
        result["success"] = True
        result["language_stats"] = language_stats
        
        # 修正日誌訊息，包含衝突數量
        total_conflicts = len(conflicts)
        if log_detail:
            status = "創建並" if result["created_new"] else ""
            if result["file_skipped"]:
                log_detail(f"JSON ({bt_code}) {status}處理完成但跳過保存：合併 {result['merged']} 個，跳過 {result['skipped']} 個，衝突 {total_conflicts} 個")
            else:
                log_detail(f"JSON ({bt_code}) {status}合併完成：合併 {result['merged']} 個，跳過 {result['skipped']} 個，衝突 {total_conflicts} 個")
        
    except json.JSONDecodeError as e:
        error_msg = f"JSON 格式錯誤：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    except Exception as e:
        error_msg = f"JSON 檔案合併失敗：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    
    return result


def handle_json_conflict(path: str, existing_value: str, new_value: str, language: str) -> str:
    """處理 JSON 合併衝突，讓用戶選擇如何處理"""
    print(f"\n⚠️  發現衝突：")
    print(f"📍 路徑：{path}")
    print(f"🌍 語言：{language}")
    print(f"📄 現有值：'{existing_value}'")
    print(f"🆕 新值：'{new_value}'")
    
    while True:
        print(f"\n請選擇處理方式：")
        print(f"  1) 保留現有值 ('{existing_value}')")
        print(f"  2) 使用新值 ('{new_value}')")
        print(f"  3) 跳過此項目")
        print(f"  A) 對所有類似衝突使用新值")
        print(f"  K) 對所有類似衝突保留現有值")
        
        try:
            choice = input(f"請選擇 (1/2/3/A/K)：").strip().upper()
            
            if choice == "1":
                return "keep_existing"
            elif choice == "2":
                return "use_new"
            elif choice == "3":
                return "skip"
            elif choice == "A":
                print(f"✅ 將使用新值")
                return "use_new"
            elif choice == "K":
                print(f"✅ 將保留現有值")
                return "keep_existing"
            else:
                print(f"⚠️  請輸入有效選項 (1/2/3/A/K)")
                
        except KeyboardInterrupt:
            print(f"\n❌ 操作取消，跳過此項目")
            return "skip"


def generate_conflict_report(conflicts: list, output_dir: Path, timestamp: str):
    """生成衝突報告"""
    if not conflicts:
        return
    
    conflict_report_file = output_dir / f"conflicts_report_{timestamp}.txt"
    
    try:
        with open(conflict_report_file, 'w', encoding='utf-8') as f:
            f.write(f"JSON 合併衝突報告\n")
            f.write(f"生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*60}\n\n")
            
            f.write(f"總衝突數量：{len(conflicts)}\n\n")
            
            for i, conflict in enumerate(conflicts, 1):
                f.write(f"衝突 {i}：\n")
                f.write(f"  路徑：{conflict['path']}\n")
                f.write(f"  語言：{conflict['language']}\n")
                f.write(f"  現有值：'{conflict['existing_value']}'\n")
                f.write(f"  新值：'{conflict['new_value']}'\n")
                f.write(f"  檔案類型：{conflict['file_type']}\n")
                f.write(f"\n{'-'*40}\n\n")
            
            f.write(f"處理建議：\n")
            f.write(f"1. 檢查值的差異是否為預期的更新\n")
            f.write(f"2. 確認語言翻譯的正確性\n")
            f.write(f"3. 驗證業態特定的術語使用\n")
            f.write(f"4. 考慮建立翻譯一致性檢查機制\n")
        
        print(f"📄 衝突報告已生成：{conflict_report_file}")
        
    except Exception as e:
        print(f"⚠️  生成衝突報告失敗：{e}")


def combine_po_files_for_business_type(all_updates: dict, target_po_path: Path, 
                                     output_dir: Path, bt_code: str, log_detail=None, 
                                     create_new: bool = False) -> dict:
    """【增強版】為特定業態處理 PO 檔案合併，每個語言生成獨立的 PO 檔案"""
    result = {
        "success": False,
        "merged": 0,
        "skipped": 0,
        "conflicts": [],
        "errors": [],
        "language_stats": {},
        "created_new": False,
        "created_files": []  # 新增：記錄創建的檔案
    }
    
    # 檢查是否有當前業態的 PO 更新
    languages_with_po_updates = {}
    for language, language_updates in all_updates.items():
        if bt_code in language_updates and language_updates[bt_code]['po']:
            languages_with_po_updates[language] = language_updates[bt_code]['po']
    
    if not languages_with_po_updates:
        result["success"] = True
        if log_detail:
            log_detail(f"PO ({bt_code}): 沒有任何更新項目")
        return result
    
    try:
        config = get_config()
        business_types = config.get_business_types()
        suffix = business_types[bt_code]['suffix'] if bt_code in business_types else ""
        
        # 為每個語言分別處理 PO 檔案
        for language, po_updates in languages_with_po_updates.items():
            print(f"   🌐 處理 {language} 的 PO 檔案...")
            if log_detail:
                log_detail(f"開始處理 {language} 的 PO 檔案 (業態: {bt_code})")
            
            # 確定當前語言的輸出檔案路徑
            if target_po_path and target_po_path != "CREATE_NEW":
                # 基於原始檔案名稱，添加語言和業態後綴
                base_name = target_po_path.stem
                # 移除可能已存在的語言後綴，避免重複
                if base_name.endswith(f"_{language}"):
                    base_name = base_name[:-len(f"_{language}")]
                output_po_path = output_dir / f"{base_name}_{language}{suffix}_combined.po"
            else:
                output_po_path = output_dir / f"messages_{language}{suffix}_combined.po"
            
            # 記錄創建的檔案
            result["created_files"].append(str(output_po_path))
            
            # 為當前語言創建或載入 PO 檔案
            if create_new or target_po_path == "CREATE_NEW" or not target_po_path or not target_po_path.exists():
                # 創建新的 PO 檔案
                print(f"     🆕 創建新的 PO 檔案：{output_po_path.name}")
                if log_detail:
                    log_detail(f"創建新的 PO 檔案：{output_po_path.name}")
                
                # 創建預設檔案
                if not create_default_po_file(output_po_path, language):
                    result["errors"].append(f"無法為 {language} 創建預設 PO 檔案")
                    continue
                
                target_po = polib.pofile(str(output_po_path))
                result["created_new"] = True
                
                # 清空預設條目，將由更新資料填充
                target_po.clear()
                
            else:
                # 嘗試載入對應語言的現有 PO 檔案
                language_specific_path = target_po_path.parent / f"{target_po_path.stem}_{language}.po"
                if language_specific_path.exists():
                    target_po = polib.pofile(str(language_specific_path))
                    print(f"     📄 載入 {language} 專用 PO 檔案：{language_specific_path.name}")
                    if log_detail:
                        log_detail(f"載入 {language} 專用 PO 檔案：{language_specific_path.name}")
                else:
                    # 使用通用 PO 檔案作為基礎
                    target_po = polib.pofile(str(target_po_path))
                    print(f"     📄 基於通用 PO 檔案創建 {language} 版本")
                    if log_detail:
                        log_detail(f"基於通用 PO 檔案創建 {language} 版本")
            
            # 初始化當前語言的統計
            language_stats = {"merged": 0, "skipped": 0, "conflicts": 0}
            
            # 處理當前語言的 PO 更新
            for msgid, new_msgstr, _ in po_updates:
                target_entry = target_po.find(msgid)
                
                if target_entry:
                    # 只有當現有值和新值真的不同時才需要更新
                    if target_entry.msgstr and target_entry.msgstr.strip():
                        if target_entry.msgstr == new_msgstr:
                            # 值相同，跳過
                            language_stats["skipped"] += 1
                            result["skipped"] += 1
                            if log_detail:
                                log_detail(f"[{language}] 跳過相同值：{msgid} = '{new_msgstr}'")
                            continue
                        else:
                            # 值不同，記錄但仍然更新
                            if log_detail:
                                log_detail(f"[{language}] 更新現有條目：{msgid} = '{new_msgstr}' (原值: '{target_entry.msgstr}')")
                    
                    # 應用更新
                    target_entry.msgstr = new_msgstr
                    language_stats["merged"] += 1
                    result["merged"] += 1
                    
                else:
                    # 目標檔案中沒有此條目，添加新條目
                    new_entry = polib.POEntry(
                        msgid=msgid,
                        msgstr=new_msgstr
                    )
                    target_po.append(new_entry)
                    language_stats["merged"] += 1
                    result["merged"] += 1
                    if log_detail:
                        log_detail(f"[{language}] 新增條目：{msgid} = '{new_msgstr}'")
            
            # 更新 PO 檔案的語言元數據
            if 'Language' in target_po.metadata:
                target_po.metadata['Language'] = language
            if 'Language-Team' in target_po.metadata:
                target_po.metadata['Language-Team'] = f'{language} <LL@li.org>'
            
            # 保存當前語言的 PO 檔案
            output_po_path.parent.mkdir(parents=True, exist_ok=True)
            target_po.save(str(output_po_path))
            
            # 記錄語言統計
            result["language_stats"][language] = language_stats
            
            print(f"     ✅ {language}: 合併 {language_stats['merged']} 個，跳過 {language_stats['skipped']} 個")
            if log_detail:
                log_detail(f"[{language}] PO 檔案處理完成：合併 {language_stats['merged']} 個，跳過 {language_stats['skipped']} 個")
        
        result["success"] = True
        
        if log_detail:
            log_detail(f"PO ({bt_code}) 處理完成：共處理 {len(languages_with_po_updates)} 個語言")
            log_detail(f"總計：合併 {result['merged']} 個，跳過 {result['skipped']} 個")
            log_detail(f"生成檔案：{', '.join(result['created_files'])}")
        
    except Exception as e:
        error_msg = f"PO 檔案合併失敗：{e}"
        result["errors"].append(error_msg)
        if log_detail:
            log_detail(f"錯誤：{error_msg}")
    
    return result


def get_json_value_by_path(data: dict, path: str):
    """按路徑獲取 JSON 值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for part_type, part_value in path_parts:
            if part_type == 'key':
                if part_value not in current:
                    return None
                current = current[part_value]
            elif part_type == 'index':
                if not isinstance(current, list) or len(current) <= part_value:
                    return None
                current = current[part_value]
        
        return current
        
    except Exception:
        return None


def set_json_value_by_path(data: dict, path: str, new_value) -> bool:
    """【v1.6 增強版】按路徑設置 JSON 值，支援陣列和普通值"""
    try:
        path_parts = parse_json_path(path)
        current = data
        
        for i, (part_type, part_value) in enumerate(path_parts):
            is_last = (i == len(path_parts) - 1)
            
            if part_type == 'key':
                if is_last:
                    current[part_value] = new_value
                else:
                    if part_value not in current:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
            
            elif part_type == 'index':
                if is_last:
                    while len(current) <= part_value:
                        current.append(None)
                    current[part_value] = new_value
                else:
                    while len(current) <= part_value:
                        current.append(None)
                    if current[part_value] is None:
                        next_part_type = path_parts[i + 1][0] if i + 1 < len(path_parts) else 'key'
                        current[part_value] = [] if next_part_type == 'index' else {}
                    current = current[part_value]
        
        return True
        
    except Exception as e:
        print(f"⚠️  設置JSON路徑失敗：{path} = {new_value}, 錯誤：{e}")
        return False


def parse_json_path(path: str) -> list:
    """解析 JSON 路徑"""
    parts = []
    current = ""
    in_bracket = False
    
    for char in path:
        if char == '[':
            if current:
                parts.append(('key', current))
                current = ""
            in_bracket = True
        elif char == ']':
            if in_bracket and current:
                try:
                    parts.append(('index', int(current)))
                except ValueError:
                    raise ValueError(f"無效的陣列索引：{current}")
                current = ""
            in_bracket = False
        elif char == '.' and not in_bracket:
            if current:
                parts.append(('key', current))
                current = ""
        else:
            current += char
    
    if current:
        parts.append(('key', current))
    
    return parts


def check_po_updates_exist(all_updates: dict) -> bool:
    """檢查是否存在任何 PO 更新"""
    for language_updates in all_updates.values():
        for bt_code, bt_updates in language_updates.items():
            if bt_updates['po']:
                return True
    return False


def check_json_updates_exist(all_updates: dict) -> bool:
    """檢查是否存在任何 JSON 更新"""
    for language_updates in all_updates.values():
        for bt_code, bt_updates in language_updates.items():
            if bt_updates['json']:
                return True
    return False


def generate_multilang_summary_report(results: dict, all_updates: dict, output_dir: Path, timestamp: str, 
                                     target_json_path: Path, target_po_path: Path, log_detail):
    """生成多語言合併處理摘要報告"""
    summary_file = output_dir / f"multi_combine_summary_{timestamp}.txt"
    
    try:
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write(f"多語言檔案合併處理摘要報告\n")
            f.write(f"生成時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*60}\n\n")
            
            f.write(f"目標檔案：\n")
            if target_json_path:
                if target_json_path == "CREATE_NEW":
                    f.write(f"  JSON: 創建新的多語言 JSON\n")
                else:
                    f.write(f"  JSON: {target_json_path}\n")
            if target_po_path:
                if target_po_path == "CREATE_NEW":
                    f.write(f"  PO: 創建新的 messages.po\n")
                else:
                    f.write(f"  PO: {target_po_path}\n")
            f.write(f"\n")
            
            f.write(f"處理的語言：\n")
            for language in all_updates.keys():
                f.write(f"  - {language}\n")
            f.write(f"\n")
            
            total_merged = 0
            total_skipped = 0
            total_errors = 0
            successful_business_types = []
            failed_business_types = []
            created_new_files = []
            skipped_files = []  # 【v1.7 新增】跳過的檔案
            
            # 按業態統計
            for bt_code, bt_results in results.items():
                f.write(f"業態：{bt_code}\n")
                
                bt_merged = sum(result.get('merged', 0) for result in bt_results.values())
                bt_skipped = sum(result.get('skipped', 0) for result in bt_results.values())
                bt_errors = []
                bt_new_files = []
                bt_skipped_files = []  # 【v1.7 新增】
                
                for result_key, result in bt_results.items():
                    bt_errors.extend(result.get('errors', []))
                    if result.get('created_new'):
                        file_type = "JSON檔案" if "json" in result_key else "PO檔案"
                        bt_new_files.append(file_type)
                    if result.get('file_skipped'):  # 【v1.7 新增】
                        file_type = "JSON檔案" if "json" in result_key else "PO檔案"
                        bt_skipped_files.append(file_type)
                
                f.write(f"合併數量：{bt_merged}\n")
                f.write(f"跳過數量：{bt_skipped}\n")
                
                if bt_new_files:
                    f.write(f"新建檔案：{', '.join(bt_new_files)}\n")
                    created_new_files.extend(bt_new_files)
                
                # 【v1.7 新增】跳過檔案統計
                if bt_skipped_files:
                    f.write(f"跳過檔案：{', '.join(bt_skipped_files)} (無實際內容)\n")
                    skipped_files.extend(bt_skipped_files)
                
                # 語言級別統計
                f.write(f"語言統計：\n")
                for result in bt_results.values():
                    if 'language_stats' in result:
                        for lang, stats in result['language_stats'].items():
                            f.write(f"  {lang}: 合併 {stats['merged']}, 跳過 {stats['skipped']}, 衝突 {stats.get('conflicts', 0)}\n")
                
                if bt_errors:
                    f.write(f"錯誤：\n")
                    for error in bt_errors:
                        f.write(f"  - {error}\n")
                    failed_business_types.append(bt_code)
                else:
                    successful_business_types.append(bt_code)
                
                total_merged += bt_merged
                total_skipped += bt_skipped
                total_errors += len(bt_errors)
                
                f.write(f"\n{'-'*40}\n\n")
            
            # 總計統計
            f.write(f"處理總結：\n")
            f.write(f"成功業態：{len(successful_business_types)}\n")
            f.write(f"失敗業態：{len(failed_business_types)}\n")
            f.write(f"總合併項目：{total_merged}\n")
            f.write(f"總跳過項目：{total_skipped}\n")
            f.write(f"總錯誤項目：{total_errors}\n")
            f.write(f"處理語言數：{len(all_updates)}\n")
            
            if created_new_files:
                f.write(f"新建檔案數：{len(set(created_new_files))}\n")
            
            # 【v1.7 新增】跳過檔案統計
            if skipped_files:
                f.write(f"跳過檔案數：{len(set(skipped_files))} (因無實際內容)\n")
            
            if successful_business_types:
                f.write(f"\n成功的業態：{', '.join(successful_business_types)}\n")
            
            if failed_business_types:
                f.write(f"失敗的業態：{', '.join(failed_business_types)}\n")
            
            # v1.7 版本新增說明
            f.write(f"\n多語言合併說明：\n")
            f.write(f"- 本次處理支援多個語言的 tobemodified 合併\n")
            f.write(f"- JSON 檔案：採用多語言結構，所有語言合併到同一檔案\n")
            f.write(f"- PO 檔案：每個語言生成獨立的 PO 檔案（如 messages_zh_TW_rt.po）\n")
            f.write(f"- 自動檢測並處理語言層級的路徑映射\n")
            f.write(f"- 按業態分別處理，避免業態間相互干擾\n")
            f.write(f"- 相同 key 且相同 value 的項目會自動跳過\n")
            f.write(f"- 不同 value 的項目會正常更新\n")
            f.write(f"- 沒有目標檔案時會自動創建標準檔案（JSON/PO）\n")
            
            f.write(f"\nv1.6 版本功能 - 智能陣列處理：\n")
            f.write(f"- 檢測陣列索引路徑（如 slogan[1]）並自動進行完整陣列更新\n")
            f.write(f"- 從 i18n_input/{{language}}/{{language}}.json 讀取原始完整陣列\n")
            f.write(f"- 只替換指定索引的元素，保持其他元素不變\n")
            f.write(f"- 避免陣列部分更新導致其他位置變成 null 的問題\n")
            f.write(f"- 支援嵌套陣列路徑（如 data.items[0].tags[2]）\n")
            f.write(f"- 當無法獲取原始陣列時，自動降級為傳統索引更新\n")
            f.write(f"- 非陣列索引路徑仍使用原有的更新邏輯\n")
            
            # 【v1.7 新增】智能檔案生成說明
            f.write(f"\nv1.7 版本新增功能 - 智能檔案生成：\n")
            f.write(f"- 自動檢測 JSON 檔案是否包含實際內容\n")
            f.write(f"- 只有在有實際更新或有意義的新內容時才生成 JSON 檔案\n")
            f.write(f"- 避免生成空的或僅包含空結構的 JSON 檔案\n")
            f.write(f"- PO 檔案仍按原有邏輯處理，不受此功能影響\n")
            f.write(f"- 提供清晰的跳過檔案統計和日誌記錄\n")
            
            f.write(f"\n使用建議：\n")
            f.write(f"- 確認目標 JSON 檔案採用多語言結構（頂層為語言代碼）\n")
            f.write(f"- PO 檔案會按語言分別生成，便於獨立管理各語言翻譯\n")
            f.write(f"- 合併前建議備份原始檔案\n")
            f.write(f"- 合併後請測試翻譯檔案的正確性\n")
            f.write(f"- 檢查各語言檔案的數據完整性\n")
            f.write(f"- 新建的檔案包含標準結構，無預設範例\n")
            f.write(f"- 確保 i18n_input 目錄包含各語言的原始 JSON 檔案以支援陣列更新\n")
            f.write(f"- 陣列索引更新會自動從原始檔案讀取完整陣列進行智能合併\n")
            f.write(f"- 如果 JSON 檔案被跳過，表示該業態沒有實際的更新內容\n")
            
        log_detail(f"多語言合併摘要報告已生成：{summary_file}")
        
    except Exception as e:
        log_detail(f"生成多語言合併摘要報告失敗：{e}")


def main():
    """主執行函數"""
    print("🚀 開始多語言檔案合併處理 (v1.7 - 修復空檔案生成版)")
    
    # 載入配置
    config = get_config()
    
    # 檢測可用的 tobemodified 檔案
    available_files = detect_tobemodified_files(config)
    
    if not available_files:
        print("❌ 未找到任何 tobemodified 檔案")
        print("請先執行 script_01_generate_xlsx.py 生成檔案")
        sys.exit(1)
    
    # 步驟1：選擇 tobemodified 檔案（支援多選）
    selected_files = choose_tobemodified_files(available_files)
    if not selected_files:
        sys.exit(1)
    
    # 檢查 i18n_combine 目錄
    combine_dir = Path("i18n_combine")
    
    if not combine_dir.exists():
        print(f"❌ 合併目錄不存在：{combine_dir}")
        print(f"請創建目錄並放入要合併的檔案")
        sys.exit(1)
    
    print(f"📁 掃描合併目錄：{combine_dir}")
    
    # 掃描 combine 目錄中的檔案
    combine_files = scan_combine_directory(combine_dir)
    
    # 步驟2：選擇要合併的 JSON 檔案
    target_json_path = choose_combine_file(combine_files['json'], 'json')
    
    # 步驟3：選擇要合併的 PO 檔案
    target_po_path = choose_combine_file(combine_files['po'], 'po')
    
    # 讀取所有選中語言的 Excel 更新資料
    all_updates = {}
    detected_languages = []
    
    for language, xlsx_path in selected_files.items():
        updates = read_excel_updates_for_language(xlsx_path, language, config)
        if updates:
            all_updates[language] = updates
            detected_languages.append(language)
    
    if not all_updates:
        print("❌ 沒有讀取到任何有效的更新資料")
        sys.exit(1)
    
    # 檢查是否有更新
    has_json_updates = check_json_updates_exist(all_updates)
    has_po_updates = check_po_updates_exist(all_updates)
    
    # 如果沒有選擇 JSON 檔案但有 JSON 更新，詢問是否創建新檔案
    if not target_json_path and has_json_updates:
        print(f"\n💡 檢測到 JSON 更新但未選擇目標檔案")
        while True:
            try:
                choice = input(f"是否創建新的多語言 JSON 檔案？(Y/n)：").strip().lower()
                if choice in ['', 'y', 'yes']:
                    target_json_path = "CREATE_NEW"
                    print(f"✅ 將創建新的多語言 JSON 檔案")
                    break
                elif choice in ['n', 'no']:
                    print(f"⏭️  跳過 JSON 檔案處理")
                    break
                else:
                    print(f"⚠️  請輸入 Y 或 N")
            except KeyboardInterrupt:
                print(f"\n❌ 操作取消")
                target_json_path = None
                break
    
    # 如果沒有選擇 PO 檔案但有 PO 更新，詢問是否創建新檔案
    if not target_po_path and has_po_updates:
        print(f"\n💡 檢測到 PO 更新但未選擇目標檔案")
        while True:
            try:
                choice = input(f"是否創建新的 messages.po 檔案？(Y/n)：").strip().lower()
                if choice in ['', 'y', 'yes']:
                    target_po_path = "CREATE_NEW"
                    print(f"✅ 將創建新的 messages.po 檔案")
                    break
                elif choice in ['n', 'no']:
                    print(f"⏭️  跳過 PO 檔案處理")
                    break
                else:
                    print(f"⚠️  請輸入 Y 或 N")
            except KeyboardInterrupt:
                print(f"\n❌ 操作取消")
                target_po_path = None
                break
    
    # 檢查是否至少選擇了一個檔案或有更新需要處理
    if not target_json_path and not target_po_path:
        print("❌ 必須至少選擇一個檔案進行合併")
        sys.exit(1)
    
    # 統計所有業態
    all_business_types = set()
    for language_updates in all_updates.values():
        all_business_types.update(language_updates.keys())
    
    print(f"\n📋 合併設定：")
    print(f"   來源語言：{', '.join(selected_files.keys())}")
    if target_json_path:
        if target_json_path == "CREATE_NEW":
            print(f"   JSON 檔案：將創建新的多語言 JSON")
        else:
            print(f"   JSON 檔案：{target_json_path.relative_to(combine_dir)}")
    if target_po_path:
        if target_po_path == "CREATE_NEW":
            print(f"   PO 檔案：將創建新的 messages.po")
        else:
            print(f"   PO 檔案：{target_po_path.relative_to(combine_dir)}")
    print(f"   涵蓋業態：{', '.join([config.get_business_types()[bt]['display_name'] for bt in all_business_types])}")
    
    # 顯示陣列更新功能提示
    print(f"\n🔧 v1.7 新功能：智能檔案生成 + 陣列處理")
    print(f"   - 自動檢測陣列索引路徑（如 slogan[1]）")
    print(f"   - 從 i18n_input/{{language}}/{{language}}.json 讀取原始陣列")
    print(f"   - 進行完整陣列更新，避免其他位置變成 null")
    print(f"   - 只在有實際內容時才生成 JSON 檔案，避免空檔案")
    
    # 檢查 i18n_input 目錄
    input_dir = Path("i18n_input")
    if not input_dir.exists():
        print(f"⚠️  未找到 i18n_input 目錄，陣列更新功能可能受限")
    else:
        missing_languages = []
        for language in detected_languages:
            language_file = input_dir / language / f"{language}.json"
            if not language_file.exists():
                missing_languages.append(language)
        
        if missing_languages:
            print(f"⚠️  缺少原始語言檔案：{', '.join(missing_languages)}")
            print(f"   陣列更新將降級為傳統索引更新")
        else:
            print(f"✅ 所有語言的原始檔案都已找到，支援完整陣列更新")
    
    # 建立輸出目錄
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    dirs = config.get_directories()
    output_dir = Path(dirs['output_dir']) / f"multi_{timestamp}_combined"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 設置日誌
    log_file = output_dir / f"multi_combine_{timestamp}.log"
    
    def log_detail(message: str):
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now().strftime('%H:%M:%S')} - {message}\n")
    
    log_detail(f"開始多語言合併處理 (v1.7)")
    log_detail(f"語言：{', '.join(selected_files.keys())}")
    log_detail(f"來源檔案：{list(selected_files.values())}")
    log_detail(f"涵蓋業態：{', '.join(all_business_types)}")
    log_detail(f"陣列更新功能：啟用")
    log_detail(f"智能檔案生成：啟用")
    
    # 處理合併邏輯 - 避免業態間衝突
    business_types = config.get_business_types()
    all_results = {}
    
    # 按業態分別處理，避免相互干擾
    for bt_code in all_business_types:
        if bt_code not in business_types:
            continue
            
        bt_config = business_types[bt_code]
        display_name = bt_config['display_name']
        suffix = bt_config['suffix']
        
        print(f"\n📝 處理 {display_name}...")
        log_detail(f"開始處理業態：{display_name}")
        
        results = {}
        
        # 為當前業態處理 JSON 檔案
        if target_json_path:
            if target_json_path == "CREATE_NEW":
                output_json_path = output_dir / f"multilang{suffix}_combined.json"
            else:
                output_json_path = output_dir / f"{target_json_path.stem}{suffix}_combined.json"
            
            create_new = (target_json_path == "CREATE_NEW")
            json_result = combine_multilang_json_files_for_business_type(
                all_updates,
                target_json_path if not create_new else None,
                output_json_path,
                bt_code,
                log_detail,
                create_new,
                detected_languages
            )
            results['json_result'] = json_result
            
            # 【v1.7 修改】顯示結果，包含檔案跳過情況
            if json_result.get('errors'):
                print(f"     ❌ JSON 檔案處理錯誤：{json_result['errors']}")
            elif json_result.get('file_skipped'):
                print(f"     ⚠️  {display_name}: 無實際內容，跳過生成 JSON 檔案")
            else:
                # 顯示語言統計
                if json_result.get('language_stats'):
                    for lang, stats in json_result['language_stats'].items():
                        if stats['merged'] > 0 or stats['skipped'] > 0:
                            print(f"     📊 {lang}: 合併 {stats['merged']} 個, 跳過 {stats['skipped']} 個")
                
                if json_result.get('created_new'):
                    print(f"     🆕 創建了新的 JSON 檔案")
                
                if json_result.get('merged', 0) == 0 and json_result.get('skipped', 0) == 0:
                    if not json_result.get('created_new'):
                        print(f"     ℹ️  {display_name} 沒有 JSON 更新項目")
        
        # 為當前業態處理 PO 檔案
        if target_po_path:
            create_new = (target_po_path == "CREATE_NEW")
            po_result = combine_po_files_for_business_type(
                all_updates,
                target_po_path if not create_new else None,
                output_dir,
                bt_code,
                log_detail,
                create_new
            )
            results['po_result'] = po_result
            
            # 顯示結果
            if po_result.get('errors'):
                print(f"     ❌ PO 檔案處理錯誤：{po_result['errors']}")
            else:
                # 顯示語言統計
                if po_result.get('language_stats'):
                    for lang, stats in po_result['language_stats'].items():
                        if stats['merged'] > 0 or stats['skipped'] > 0:
                            print(f"     📊 {lang}: 合併 {stats['merged']} 個, 跳過 {stats['skipped']} 個")
                
                if po_result.get('created_new'):
                    print(f"     🆕 創建了新的 PO 檔案")
                
                if po_result.get('merged', 0) == 0 and po_result.get('skipped', 0) == 0:
                    if not po_result.get('created_new'):
                        print(f"     ℹ️  {display_name} 沒有 PO 更新項目")
        
        # 【v1.7 修改】如果沒有更新，複製原檔案（僅限非創建新檔案且未跳過的情況）
        if target_json_path and target_json_path != "CREATE_NEW" and results.get('json_result', {}).get('merged', 0) == 0:
            if (not results.get('json_result', {}).get('created_new', False) and 
                not results.get('json_result', {}).get('file_skipped', False)):
                output_json_path = output_dir / f"{target_json_path.stem}{suffix}_combined.json"
                if not output_json_path.exists():
                    output_json_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(target_json_path, output_json_path)
                    print(f"     📄 複製 JSON 檔案（無更新）")
                    log_detail(f"複製原始 JSON 檔案：{target_json_path.name}")
        
        # PO 檔案現在是按語言分別生成，所以不需要複製邏輯
        
        all_results[bt_code] = results
        
        # 統計結果
        total_merged = 0
        total_skipped = 0
        total_errors = 0
        has_new_files = False
        has_skipped_files = False  # 【v1.7 新增】
        
        for result in results.values():
            total_merged += result.get('merged', 0)
            total_skipped += result.get('skipped', 0)
            total_errors += len(result.get('errors', []))
            if result.get('created_new'):
                has_new_files = True
            if result.get('file_skipped'):  # 【v1.7 新增】
                has_skipped_files = True
        
        if total_errors > 0:
            print(f"     ❌ 處理失敗 - 錯誤: {total_errors} 個")
        else:
            status_msg = f"完成 - 合併: {total_merged} 個, 跳過: {total_skipped} 個"
            if has_new_files:
                status_msg += " (包含新檔案)"
            if has_skipped_files:  # 【v1.7 新增】
                status_msg += " (部分檔案因無內容被跳過)"
            print(f"     ✅ {status_msg}")
        
        log_detail(f"{display_name} 處理完成：合併 {total_merged} 個，跳過 {total_skipped} 個，錯誤 {total_errors} 個")
    
    # 生成最終報告
    total_merged = sum(
        sum(result.get('merged', 0) for result in results.values())
        for results in all_results.values()
    )
    total_skipped = sum(
        sum(result.get('skipped', 0) for result in results.values())
        for results in all_results.values()
    )
    total_errors = sum(
        sum(len(result.get('errors', [])) for result in results.values())
        for results in all_results.values()
    )
    
    print(f"\n🎉 多語言合併處理完成！(v1.7)")
    print(f"📊 處理結果：合併 {total_merged} 個項目，跳過 {total_skipped} 個項目")
    if total_errors > 0:
        print(f"⚠️  處理錯誤：{total_errors} 個")
    print(f"📁 輸出目錄：{output_dir}")
    print(f"🔧 新功能：智能檔案生成已啟用，自動跳過無內容的 JSON 檔案")
    print(f"🔧 陣列更新功能：已啟用，自動處理陣列索引路徑")
    
    # 生成處理摘要
    generate_multilang_summary_report(all_results, all_updates, output_dir, timestamp, target_json_path, target_po_path, log_detail)


if __name__ == "__main__":
    main()