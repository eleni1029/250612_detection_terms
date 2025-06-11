#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_based_mapping.py

基於 phrase_comparison.xlsx 的安全映射方案
直接從 Excel 讀取敏感詞到方案的對應關係，避免順序依賴風險
"""

from pathlib import Path
import openpyxl
import sys
from config_loader import get_config

class ExcelBasedMapping:
    """基於 Excel 的敏感詞映射類"""
    
    def __init__(self, excel_path: str = "phrase_comparison.xlsx"):
        """
        初始化 Excel 映射
        
        Args:
            excel_path: phrase_comparison.xlsx 檔案路徑
        """
        self.excel_path = Path(excel_path)
        self.config = get_config()
        self.mappings = {}
        self.load_mappings()
    
    def load_mappings(self):
        """從 Excel 檔案載入映射關係"""
        if not self.excel_path.exists():
            print(f"❌ 找不到 {self.excel_path}")
            print("請先執行 phrase_comparison.py 生成 Excel 檔案")
            sys.exit(1)
        
        try:
            print(f"📖 從 {self.excel_path} 載入映射關係...")
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            
            # 讀取標題列
            header_row = list(ws[1])
            headers = [str(cell.value).strip() if cell.value else "" for cell in header_row]
            
            # 建立欄位索引映射
            column_indices = {header: idx for idx, header in enumerate(headers)}
            
            # 檢查必要欄位
            required_columns = ["敏感詞類型", "敏感詞"]
            missing_columns = []
            
            for col in required_columns:
                if col not in column_indices:
                    missing_columns.append(col)
            
            # 檢查業態欄位
            business_types = self.config.get_business_types()
            for bt_code, bt_config in business_types.items():
                display_name = bt_config['display_name']
                solution_col = f"對應方案({display_name})"
                if solution_col not in column_indices:
                    missing_columns.append(solution_col)
            
            if missing_columns:
                print(f"❌ Excel 缺少必要欄位：{missing_columns}")
                print(f"現有欄位：{headers}")
                sys.exit(1)
            
            # 初始化映射字典
            for bt_code in business_types.keys():
                self.mappings[bt_code] = {}
            
            # 讀取資料行
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                
                # 安全讀取欄位值
                def get_cell_value(col_name):
                    if col_name in column_indices:
                        idx = column_indices[col_name]
                        if idx < len(row) and row[idx] is not None:
                            return str(row[idx]).strip()
                    return ""
                
                category = get_cell_value("敏感詞類型")
                keyword = get_cell_value("敏感詞")
                
                if not category or not keyword:
                    continue
                
                # 讀取各業態的對應方案
                for bt_code, bt_config in business_types.items():
                    display_name = bt_config['display_name']
                    solution_col = f"對應方案({display_name})"
                    solution = get_cell_value(solution_col)
                    
                    # 如果沒有方案，使用原敏感詞
                    if not solution:
                        solution = keyword
                    
                    self.mappings[bt_code][keyword] = solution
                
                row_count += 1
            
            print(f"✅ 成功載入 {row_count} 個敏感詞的映射關係")
            
            # 顯示載入統計
            for bt_code, bt_config in business_types.items():
                display_name = bt_config['display_name']
                mapping_count = len(self.mappings[bt_code])
                replaced_count = sum(1 for k, v in self.mappings[bt_code].items() if k != v)
                print(f"   {display_name}: {mapping_count} 個敏感詞, {replaced_count} 個有替換方案")
            
        except Exception as e:
            print(f"❌ 載入 Excel 檔案失敗：{e}")
            sys.exit(1)
    
    def get_mapping(self, business_type_code: str) -> dict:
        """
        獲取指定業態的映射字典
        
        Args:
            business_type_code: 業態代碼
        
        Returns:
            敏感詞到方案的映射字典
        """
        if business_type_code not in self.mappings:
            print(f"❌ 未知的業態代碼：{business_type_code}")
            return {}
        
        return self.mappings[business_type_code]
    
    def get_all_mappings(self) -> dict:
        """獲取所有業態的映射字典"""
        return self.mappings
    
    def get_replacement(self, keyword: str, business_type_code: str) -> str:
        """
        獲取指定敏感詞在指定業態下的替換方案
        
        Args:
            keyword: 敏感詞
            business_type_code: 業態代碼
        
        Returns:
            替換方案，如果沒有則返回原敏感詞
        """
        mapping = self.get_mapping(business_type_code)
        return mapping.get(keyword, keyword)
    
    def apply_replacements(self, text: str, business_type_code: str) -> str:
        """
        對文本應用敏感詞替換
        
        Args:
            text: 原始文本
            business_type_code: 業態代碼
        
        Returns:
            替換後的文本
        """
        if not text:
            return text
        
        mapping = self.get_mapping(business_type_code)
        result = text
        
        # 按長度排序，優先替換長詞
        sorted_keywords = sorted(mapping.keys(), key=len, reverse=True)
        
        for keyword in sorted_keywords:
            replacement = mapping[keyword]
            if keyword != replacement:
                result = result.replace(keyword, replacement)
        
        return result
    
    def build_replacement_plan(self, keywords: list, business_type_code: str) -> str:
        """
        建立替換方案說明
        
        Args:
            keywords: 敏感詞列表
            business_type_code: 業態代碼
        
        Returns:
            替換方案說明字符串
        """
        mapping = self.get_mapping(business_type_code)
        replacements = []
        
        for keyword in keywords:
            replacement = mapping.get(keyword, keyword)
            if replacement != keyword:
                replacements.append(f"{keyword}→{replacement}")
        
        return "、".join(replacements)
    
    def validate_completeness(self) -> bool:
        """
        驗證映射的完整性
        確保所有敏感詞都有對應的映射
        """
        print(f"\n🔍 驗證映射完整性...")
        
        try:
            # 載入基礎敏感詞
            from detection_terms import DETECTION_TERMS
            
            all_keywords = set()
            for category, keywords in DETECTION_TERMS.items():
                all_keywords.update(keywords)
            
            business_types = self.config.get_business_types()
            all_complete = True
            
            for bt_code, bt_config in business_types.items():
                display_name = bt_config['display_name']
                mapping = self.mappings[bt_code]
                mapped_keywords = set(mapping.keys())
                
                missing_keywords = all_keywords - mapped_keywords
                extra_keywords = mapped_keywords - all_keywords
                
                if missing_keywords:
                    print(f"⚠️  {display_name}映射缺少敏感詞：{missing_keywords}")
                    all_complete = False
                
                if extra_keywords:
                    print(f"ℹ️  {display_name}映射有額外詞彙：{extra_keywords}")
            
            if all_complete:
                print("✅ 所有映射都完整")
            else:
                print("⚠️  發現不完整的映射")
            
            return all_complete
            
        except ImportError:
            print("⚠️  無法載入 detection_terms.py 進行完整性驗證")
            return True


# 便利函數
def get_excel_mapping(excel_path: str = "phrase_comparison.xlsx") -> ExcelBasedMapping:
    """
    獲取 Excel 映射實例
    
    Args:
        excel_path: Excel 檔案路徑
    
    Returns:
        ExcelBasedMapping 實例
    """
    return ExcelBasedMapping(excel_path)


def test_mapping():
    """測試映射功能"""
    print("🧪 測試 Excel 映射功能...")
    
    try:
        mapper = get_excel_mapping()
        
        # 驗證完整性
        mapper.validate_completeness()
        
        print("✅ 測試完成")
        
    except Exception as e:
        print(f"❌ 測試失敗：{e}")


if __name__ == "__main__":
    test_mapping()