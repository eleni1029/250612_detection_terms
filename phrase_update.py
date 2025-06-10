#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
phrase_update.py

根據 phrase_comparison.xlsx 重建 4 個 detection_terms py 文件，
同時備份原始檔案到 backup/，加上 timestamp。

校驗：
- 不能有缺值（敏感詞類型、敏感詞、對應方案皆不可空）
- 必須有內容，否則報錯
- 修正資料映射邏輯錯誤
"""

import os
import pathlib
import shutil
import datetime
import openpyxl
import sys
from collections import defaultdict

# ── 檔案路徑 ────────────────────────────────
BASE_DIR = pathlib.Path(__file__).parent
BACKUP_DIR = BASE_DIR / "backup"
BACKUP_DIR.mkdir(exist_ok=True)

# 四個 py 檔案
FILES = {
    "main": BASE_DIR / "detection_terms.py",
    "enterprise": BASE_DIR / "detection_terms_enterprises.py",
    "public": BASE_DIR / "detection_terms_public_sector.py",
    "edu": BASE_DIR / "detection_terms_training_institutions.py",
}

# Excel 檔案
XLSX_PATH = BASE_DIR / "phrase_comparison.xlsx"

# ── 備份檔案 ────────────────────────────────
def backup_files():
    """備份現有檔案到 backup 目錄"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    for label, path in FILES.items():
        if path.exists():
            dst = BACKUP_DIR / f"{path.stem}_{timestamp}.py"
            shutil.copy(path, dst)
            print(f"✅ 備份 {path.name} → {dst.name}")

# ── 讀取 Excel 與檢查 ──────────────────────
def read_and_validate():
    """讀取並驗證 Excel 檔案內容"""
    print(f"\n🔍 讀取 {XLSX_PATH} ...")
    if not XLSX_PATH.exists():
        print(f"❌ 找不到 {XLSX_PATH}，停止執行")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
    except Exception as e:
        print(f"❌ 無法讀取 Excel 檔案：{e}")
        sys.exit(1)

    # 檢查是否有資料
    if ws.max_row < 2:
        print("❌ Excel 檔案沒有資料列")
        sys.exit(1)

    # 讀取標題列
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    print(f"📋 Excel 標題列：{header}")
    
    # 定義所需欄位（支援不同的標題格式）
    required_columns = {
        "敏感詞類型": ["敏感詞類型", "類型", "分類"],
        "敏感詞": ["敏感詞", "關鍵詞", "詞彙"],
        "對應方案(企業)": ["對應方案(企業)", "企業方案", "企業"],
        "對應方案(公部門)": ["對應方案(公部門)", "公部門方案", "公部門"],
        "對應方案(培訓機構)": ["對應方案(培訓機構)", "培訓機構方案", "培訓機構"]
    }
    
    # 尋找欄位索引
    column_indices = {}
    for standard_name, possible_names in required_columns.items():
        found_idx = None
        for possible_name in possible_names:
            try:
                found_idx = header.index(possible_name)
                break
            except ValueError:
                continue
        
        if found_idx is None:
            print(f"❌ 找不到必要欄位 '{standard_name}'")
            print(f"   支援的欄位名稱：{possible_names}")
            print(f"   實際欄位：{header}")
            sys.exit(1)
        
        column_indices[standard_name] = found_idx

    print(f"✅ 欄位對應：{column_indices}")

    # 讀取資料
    data = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # 安全讀取每列資料
        row_data = {}
        for standard_name, col_idx in column_indices.items():
            cell_value = row[col_idx].value
            if cell_value is None:
                row_data[standard_name] = ""
            else:
                row_data[standard_name] = str(cell_value).strip()

        # 跳過完全空白的列
        if not any(row_data.values()):
            continue

        # 檢查必要欄位
        missing_fields = []
        for field, value in row_data.items():
            if not value:
                missing_fields.append(field)

        if missing_fields:
            print(f"❌ 第 {row_num} 列缺少必要資料：{missing_fields}")
            print(f"   資料內容：{row_data}")
            sys.exit(1)

        # 轉換為元組格式以保持相容性
        data.append((
            row_data["敏感詞類型"],
            row_data["敏感詞"],
            row_data["對應方案(企業)"],
            row_data["對應方案(公部門)"],
            row_data["對應方案(培訓機構)"]
        ))

    if not data:
        print("❌ Excel 無有效內容，停止執行")
        sys.exit(1)

    print(f"✅ 成功讀取 {len(data)} 筆資料")
    
    # 顯示讀取的資料以供驗證
    print("\n📊 讀取的資料摘要：")
    categories = defaultdict(int)
    for cat, kw, ent, pub, edu in data:
        categories[cat] += 1
    
    for cat, count in categories.items():
        print(f"   {cat}: {count} 筆")
    
    return data

# ── 重組成 dict ───────────────────────────
def build_terms(data, col_idx):
    """根據 col_idx 重建 dict
    col_idx: 0=企業, 1=公部門, 2=培訓機構
    """
    terms = defaultdict(set)  # 使用 set 避免重複
    for cat, kw, ent, pub, edu in data:
        val = [ent, pub, edu][col_idx]
        terms[cat].add(val)
    
    # 轉換為 list 並排序
    return {cat: sorted(list(words)) for cat, words in terms.items()}

def build_main_terms(data):
    """建立主檔案的敏感詞字典"""
    terms = defaultdict(set)  # 使用 set 避免重複
    for cat, kw, ent, pub, edu in data:
        terms[cat].add(kw)
    
    # 轉換為 list 並排序
    return {cat: sorted(list(words)) for cat, words in terms.items()}

# ── 輸出 py 檔案 ──────────────────────────
def write_py(path, terms_dict):
    """將字典寫入 Python 檔案"""
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write("# Auto-generated by phrase_update.py\n")
            f.write("# -*- coding: utf-8 -*-\n\n")
            f.write("DETECTION_TERMS = {\n")
            for cat, words in sorted(terms_dict.items()):
                f.write(f'    "{cat}": [\n')
                for w in words:
                    # 處理特殊字符
                    escaped_w = w.replace('\\', '\\\\').replace('"', '\\"')
                    f.write(f'        "{escaped_w}",\n')
                f.write("    ],\n")
            f.write("}\n")
        print(f"✅ 成功寫入 {path.name}")
    except Exception as e:
        print(f"❌ 寫入檔案 {path} 失敗：{e}")
        sys.exit(1)

# ── 檢查分類完整性 ──────────────────────
def validate_categories(data):
    """檢查是否涵蓋全部分類，且每一類至少一詞"""
    try:
        main_file = FILES["main"]
        if not main_file.exists():
            print(f"⚠️  {main_file.name} 不存在，跳過分類完整性檢查")
            return
            
        # 動態導入模組
        import importlib.util
        spec = importlib.util.spec_from_file_location("detection_terms", main_file)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        
        BASE_TERMS = module.DETECTION_TERMS
        base_cats = set(BASE_TERMS.keys())
    except Exception as e:
        print(f"⚠️  無法讀取現有分類進行比對：{e}")
        print("    將跳過分類完整性檢查")
        return

    found_cats = set(cat for cat, *_ in data)
    missing_cats = base_cats - found_cats
    new_cats = found_cats - base_cats

    if missing_cats:
        print(f"⚠️  Excel 缺少原有分類：{missing_cats}")
    
    if new_cats:
        print(f"ℹ️  Excel 新增分類：{new_cats}")

    # 檢查每一分類至少有一詞
    counts = defaultdict(int)
    for cat, *_ in data:
        counts[cat] += 1

    empty_cats = [cat for cat, cnt in counts.items() if cnt == 0]
    if empty_cats:
        print(f"❌ 檢查失敗！以下分類沒有任何詞：{empty_cats}")
        sys.exit(1)

    print(f"✅ 分類檢查完成，Excel 共 {len(found_cats)} 類別")

# ── 資料驗證和預覽 ──────────────────────
def preview_generated_data(main_terms, ent_terms, pub_terms, edu_terms):
    """預覽將要生成的資料"""
    print("\n📋 生成資料預覽：")
    
    print("\n1. detection_terms.py (敏感詞):")
    for cat, words in list(main_terms.items())[:3]:  # 只顯示前3個類別
        print(f"   {cat}: {words}")
    if len(main_terms) > 3:
        print(f"   ... 另外 {len(main_terms) - 3} 個類別")
    
    print("\n2. detection_terms_enterprises.py (企業方案):")
    for cat, words in list(ent_terms.items())[:3]:
        print(f"   {cat}: {words}")
    if len(ent_terms) > 3:
        print(f"   ... 另外 {len(ent_terms) - 3} 個類別")
    
    print("\n3. detection_terms_public_sector.py (公部門方案):")
    for cat, words in list(pub_terms.items())[:3]:
        print(f"   {cat}: {words}")
    if len(pub_terms) > 3:
        print(f"   ... 另外 {len(pub_terms) - 3} 個類別")
    
    print("\n4. detection_terms_training_institutions.py (培訓機構方案):")
    for cat, words in list(edu_terms.items())[:3]:
        print(f"   {cat}: {words}")
    if len(edu_terms) > 3:
        print(f"   ... 另外 {len(edu_terms) - 3} 個類別")

# ── 主流程 ────────────────────────────────
def main():
    """主要執行流程"""
    print("🚀 開始執行 phrase_update.py")
    
    # 備份現有檔案
    backup_files()

    # 讀取並驗證資料
    data = read_and_validate()
    
    # 驗證分類完整性
    validate_categories(data)
    
    # 重建四份 dict
    print("\n📝 重建字典...")
    
    # 修正：使用專門的函數建立各類字典
    main_terms = build_main_terms(data)  # 敏感詞分類
    ent_terms = build_terms(data, 0)     # 企業方案
    pub_terms = build_terms(data, 1)     # 公部門方案
    edu_terms = build_terms(data, 2)     # 培訓機構方案

    # 預覽生成的資料
    preview_generated_data(main_terms, ent_terms, pub_terms, edu_terms)
    
    # 確認是否繼續
    confirm = input("\n❓ 確認要寫入檔案嗎？ (y/N): ").strip().lower()
    if confirm not in ['y', 'yes']:
        print("❌ 使用者取消操作")
        sys.exit(0)

    # 輸出檔案
    print("\n💾 寫入檔案...")
    write_py(FILES["main"], main_terms)
    write_py(FILES["enterprise"], ent_terms)
    write_py(FILES["public"], pub_terms)
    write_py(FILES["edu"], edu_terms)

    # 完成報告
    print("\n🎉 全部更新完成！統計資料：")
    print(f"  detection_terms.py → {len(main_terms)} 類別，{sum(len(words) for words in main_terms.values())} 個敏感詞")
    print(f"  detection_terms_enterprises.py → {len(ent_terms)} 類別，{sum(len(words) for words in ent_terms.values())} 個方案")
    print(f"  detection_terms_public_sector.py → {len(pub_terms)} 類別，{sum(len(words) for words in pub_terms.values())} 個方案")
    print(f"  detection_terms_training_institutions.py → {len(edu_terms)} 類別，{sum(len(words) for words in edu_terms.values())} 個方案")

if __name__ == "__main__":
    main()